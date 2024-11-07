import requests
from bs4 import BeautifulSoup
import json
from time import sleep
import random
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# User agent
user_agents = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36 OPR/113.0.0.0',
    # Adicione mais User-Agents aqui
]

# Selecionar produto
produto = input('Digite o produto: ')
produto = produto.replace(' ', '-')

# URL base
url = f'https://lista.mercadolivre.com.br/{produto}_Desde_'

# Contagem
start = 1

# Lista para armazenar os dados
dados = []

headers = {
    'User-Agent': random.choice(user_agents)
}

url_final = url + str(start) + '_NoIndex_True'
# Fazer requisição
r = requests.get(url_final, headers=headers)
site = BeautifulSoup(r.content, 'html.parser')

# Encontrar os resultados
descricoes = site.find_all('h2', class_='poly-component__title')
precos = site.find_all('span', class_='andes-money-amount andes-money-amount--cents-superscript')
links = site.find_all('li', class_='ui-search-layout__item')

# Limitando a 10 itens
for i, (descricao, preco, link) in enumerate(zip(descricoes, precos, links)):
    if i >= 10:  # Parar após 10 itens
        break
    print(i + 1)  # Imprime o número do item
    print(descricao.get_text())
    print(preco.get_text())

    a_tag = link.find('a', href=True)  # Encontrar a tag <a> dentro do <li>
    if a_tag:
        href = a_tag.get("href")  # Pega o valor do href dentro da <a>
        print(f'Link: {href}')

    # pagina produto
    r2 = requests.get(href, headers=headers)
    site2 = BeautifulSoup(r2.content, 'html.parser')

    vendidos = site2.find_all('span', class_='ui-pdp-subtitle')
    trat_vend = ''
    for vendido in vendidos:
        trat_vend = vendido.get_text().replace('Novo', '').replace('vendidos', '').replace(' ', '').replace('+', '').replace('|', '').replace('mil', '000')

    print(trat_vend)

    vendedores = site2.find_all('div', class_='ui-seller-data-header__title-container')
    trat_vendedor = ''
    for vendedor in vendedores:
        trat_vendedor = vendedor.get_text().replace('Vendido por ', '').strip()
        print(f'{trat_vendedor}\n')

    # Adicionar os dados a lista
    produto_dados = {
        'Sequencia': i + 1,
        'Descricao': descricao.get_text(),
        'Preço': preco.get_text(),
        'Vendido + de:': trat_vend,
        'Vendedor': trat_vendedor,
        'Link': href,
    }
    
    dados.append(produto_dados)

    sleep(random.randint(1, 3))  # Intervalo reduzido entre as requisições para não sobrecarregar

# Salvar dados em um arquivo JSON
with open('produtos.json', 'w', encoding='utf-8') as f:
    json.dump(dados, f, ensure_ascii=False, indent=4)
print("Dados salvos em produtos.json")

# Passo 1: Ler o arquivo JSON
json_file_path = 'produtos.json'
df = pd.read_json(json_file_path)

# Passo 2: Selecionar os dados
df_primeiros_dez = df.head(4)  # Alterado para pegar os 10 primeiros itens

# Passo 3: Salvar os dados em uma aba específica de um arquivo Excel
excel_file_path = 'raspagem.xlsx'

# Tenta abrir o arquivo Excel
try:
    wb = load_workbook(excel_file_path)
    with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_primeiros_dez.to_excel(writer, sheet_name='Dados', index=False)

except FileNotFoundError:
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        df_primeiros_dez.to_excel(writer, sheet_name='Dados', index=False)

print("Dados salvos em raspagem.xlsx")